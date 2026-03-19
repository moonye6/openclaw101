import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

with open('axing_articles.json', 'r') as f:
    all_articles = json.load(f)
with open('axing_openclaw.json', 'r') as f:
    openclaw_articles = json.load(f)
with open('axing_mcp.json', 'r') as f:
    mcp_articles = json.load(f)

axing_all = [a for a in all_articles if a['account'] == '跟阿星一起学AI']
axing_openclaw = [a for a in openclaw_articles if a['account'] == '跟阿星一起学AI']
axing_mcp = [a for a in mcp_articles if a['account'] == '跟阿星一起学AI']

seen_titles = set()
all_axing = []
for a in axing_all + axing_openclaw + axing_mcp:
    if a['title'] not in seen_titles:
        seen_titles.add(a['title'])
        all_axing.append(a)

for a in all_axing:
    a['datetime'] = datetime.strptime(a['time'], '%Y-%m-%d %H:%M')
all_axing.sort(key=lambda x: x['datetime'], reverse=True)

week_start = datetime(2026, 3, 12)
week_end = datetime(2026, 3, 19, 23, 59, 59)
recent_week = [a for a in all_axing if week_start <= a['datetime'] <= week_end]
recent_month = [a for a in all_axing if a['datetime'] >= datetime(2026, 2, 19)]

header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
header_fill = PatternFill('solid', fgColor='2563EB')
sub_header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
normal_font = Font(name='Arial', size=11)
bold_font = Font(name='Arial', bold=True, size=11)
green_fill = PatternFill('solid', fgColor='DCFCE7')
yellow_fill = PatternFill('solid', fgColor='FEF3C7')
red_fill = PatternFill('solid', fgColor='FEE2E2')
blue_fill = PatternFill('solid', fgColor='DBEAFE')
light_gray_fill = PatternFill('solid', fgColor='F8FAFC')
white_fill = PatternFill('solid', fgColor='FFFFFF')
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

weekday_map = {0: '周一', 1: '周二', 2: '周三', 3: '周四', 4: '周五', 5: '周六', 6: '周日'}

def classify_topic(title, digest):
    t = (title + ' ' + digest).lower()
    if 'openclaw' in t or '龙虾' in t:
        return 'OpenClaw/AI Agent'
    elif 'mcp' in t:
        return 'MCP 协议'
    elif 'skill' in t or 'rag' in t or '新词' in t:
        return 'AI 概念科普'
    elif '安全' in t or '裸奔' in t:
        return 'AI 安全'
    elif '视频' in t:
        return 'AI 视频创作'
    elif '日报' in t or '助手' in t:
        return 'AI 实战教程'
    elif '春节' in t or '中国' in t:
        return 'AI 行业观察'
    elif 'flowith' in t:
        return 'AI 工具评测'
    elif '元宝' in t or '腾讯' in t:
        return '产品体验'
    else:
        return 'AI 综合'

wb = Workbook()

# ========== Sheet 1: 数据概览 ==========
ws1 = wb.active
ws1.title = '数据概览'
ws1.sheet_properties.tabColor = '2563EB'

ws1.merge_cells('A1:H1')
ws1['A1'] = '「跟阿星一起学AI」公众号数据分析报告'
ws1['A1'].font = Font(name='Arial', bold=True, size=18, color='1E40AF')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 45

ws1.merge_cells('A2:H2')
ws1['A2'] = f'分析周期: 2026-03-12 ~ 2026-03-19 (最近一周) | 生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
ws1['A2'].font = Font(name='Arial', size=10, color='64748B')
ws1['A2'].alignment = Alignment(horizontal='center')

# KPI row
row = 4
kpi_data = [
    ('本周发文数', f'{len(recent_week)} 篇', blue_fill),
    ('近30天发文', f'{len(recent_month)} 篇', green_fill),
    ('周均发文频率', f'{len(recent_week)/7:.1f} 篇/天', yellow_fill),
    ('本周核心话题', 'AI Agent / 安全', red_fill),
]
for i, (label, value, fill) in enumerate(kpi_data):
    col = i * 2 + 1
    ws1.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col+1)
    cell = ws1.cell(row=row, column=col)
    cell.value = value
    cell.font = Font(name='Arial', size=18, bold=True, color='1E40AF')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    for r in range(row, row+2):
        for c in range(col, col+2):
            ws1.cell(row=r, column=c).fill = fill
            ws1.cell(row=r, column=c).border = thin_border
    ws1.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    lc = ws1.cell(row=row+2, column=col)
    lc.value = label
    lc.font = Font(name='Arial', size=10, color='475569', bold=True)
    lc.alignment = Alignment(horizontal='center')

# Article table
row = 9
ws1.merge_cells(f'A{row}:H{row}')
ws1[f'A{row}'] = '本周文章列表 (3月12日-3月19日)'
ws1[f'A{row}'].font = Font(name='Arial', bold=True, size=14, color='1E40AF')
ws1.row_dimensions[row].height = 30

row += 1
headers = ['序号', '发布日期', '发布时间', '文章标题', '内容摘要', '主题分类', '星期', '距今天数']
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

today = datetime(2026, 3, 19)
for idx, a in enumerate(recent_week, 1):
    row += 1
    dt = a['datetime']
    days_ago = (today - dt).days
    topic = classify_topic(a['title'], a.get('digest', ''))
    digest = a.get('digest', '')
    if len(digest) > 80:
        digest = digest[:80] + '...'
    data = [idx, dt.strftime('%Y-%m-%d'), dt.strftime('%H:%M'), a['title'], digest, topic, weekday_map[dt.weekday()], days_ago]
    fill = light_gray_fill if idx % 2 == 0 else white_fill
    for i, v in enumerate(data, 1):
        cell = ws1.cell(row=row, column=i, value=v)
        cell.font = normal_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.fill = fill
        cell.border = thin_border

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 10

# ========== Sheet 2: 内容分析 ==========
ws2 = wb.create_sheet('内容分析')
ws2.sheet_properties.tabColor = '10B981'

ws2.merge_cells('A1:F1')
ws2['A1'] = '内容主题与发布趋势分析'
ws2['A1'].font = Font(name='Arial', bold=True, size=16, color='065F46')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 40

# Topic distribution
row = 3
ws2[f'A{row}'] = '主题分类分布'
ws2[f'A{row}'].font = Font(name='Arial', bold=True, size=13, color='065F46')

row = 4
topic_counts = {}
for a in all_axing:
    topic = classify_topic(a['title'], a.get('digest', ''))
    topic_counts[topic] = topic_counts.get(topic, 0) + 1

total = sum(topic_counts.values())
for i, h in enumerate(['主题分类', '文章数量', '占比'], 1):
    cell = ws2.cell(row=row, column=i, value=h)
    cell.font = sub_header_font
    cell.fill = PatternFill('solid', fgColor='10B981')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for idx, (topic, count) in enumerate(sorted(topic_counts.items(), key=lambda x: -x[1]), 1):
    r = row + idx
    ws2.cell(row=r, column=1, value=topic).font = normal_font
    ws2.cell(row=r, column=2, value=count).font = bold_font
    ws2.cell(row=r, column=3, value=count/total).font = normal_font
    ws2.cell(row=r, column=3).number_format = '0.0%'
    for c in range(1, 4):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')

pie = PieChart()
pie.title = '文章主题分布'
pie.style = 10
data_ref = Reference(ws2, min_col=2, min_row=row, max_row=row+len(topic_counts))
cats = Reference(ws2, min_col=1, min_row=row+1, max_row=row+len(topic_counts))
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats)
pie.width = 18
pie.height = 12
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
ws2.add_chart(pie, f'E{row}')

# Time slot analysis
time_row = row + len(topic_counts) + 18
ws2[f'A{time_row}'] = '发布时间段分析'
ws2[f'A{time_row}'].font = Font(name='Arial', bold=True, size=13, color='065F46')
time_row += 1

time_slots = {'早间(6-10点)': 0, '午间(10-14点)': 0, '下午(14-18点)': 0, '晚间(18-22点)': 0, '深夜(22-6点)': 0}
for a in all_axing:
    h = a['datetime'].hour
    if 6 <= h < 10: time_slots['早间(6-10点)'] += 1
    elif 10 <= h < 14: time_slots['午间(10-14点)'] += 1
    elif 14 <= h < 18: time_slots['下午(14-18点)'] += 1
    elif 18 <= h < 22: time_slots['晚间(18-22点)'] += 1
    else: time_slots['深夜(22-6点)'] += 1

for i, h in enumerate(['时间段', '文章数量', '占比'], 1):
    cell = ws2.cell(row=time_row, column=i, value=h)
    cell.font = sub_header_font
    cell.fill = PatternFill('solid', fgColor='10B981')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for idx, (slot, count) in enumerate(time_slots.items(), 1):
    r = time_row + idx
    ws2.cell(row=r, column=1, value=slot).font = normal_font
    ws2.cell(row=r, column=2, value=count).font = bold_font
    ws2.cell(row=r, column=3, value=count/total).font = normal_font
    ws2.cell(row=r, column=3).number_format = '0.0%'
    for c in range(1, 4):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')

bar = BarChart()
bar.type = 'col'
bar.title = '发布时间段分布'
bar.style = 10
bar.y_axis.title = '文章数量'
data_ref2 = Reference(ws2, min_col=2, min_row=time_row, max_row=time_row+5)
cats2 = Reference(ws2, min_col=1, min_row=time_row+1, max_row=time_row+5)
bar.add_data(data_ref2, titles_from_data=True)
bar.set_categories(cats2)
bar.width = 18
bar.height = 12
ws2.add_chart(bar, f'E{time_row}')

# Weekday analysis
wd_row = time_row + 20
ws2[f'A{wd_row}'] = '星期发布分布'
ws2[f'A{wd_row}'].font = Font(name='Arial', bold=True, size=13, color='065F46')
wd_row += 1

weekday_counts = {i: 0 for i in range(7)}
for a in all_axing:
    weekday_counts[a['datetime'].weekday()] += 1

for i, h in enumerate(['星期', '文章数量'], 1):
    cell = ws2.cell(row=wd_row, column=i, value=h)
    cell.font = sub_header_font
    cell.fill = PatternFill('solid', fgColor='10B981')
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

for idx in range(7):
    r = wd_row + idx + 1
    ws2.cell(row=r, column=1, value=weekday_map[idx]).font = normal_font
    ws2.cell(row=r, column=2, value=weekday_counts[idx]).font = bold_font
    for c in range(1, 3):
        ws2.cell(row=r, column=c).border = thin_border
        ws2.cell(row=r, column=c).alignment = Alignment(horizontal='center')

bar2 = BarChart()
bar2.type = 'col'
bar2.title = '每周各天发布量'
bar2.style = 10
data_ref3 = Reference(ws2, min_col=2, min_row=wd_row, max_row=wd_row+7)
cats3 = Reference(ws2, min_col=1, min_row=wd_row+1, max_row=wd_row+7)
bar2.add_data(data_ref3, titles_from_data=True)
bar2.set_categories(cats3)
bar2.width = 18
bar2.height = 12
ws2.add_chart(bar2, f'E{wd_row}')

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 12
ws2.column_dimensions['C'].width = 10

# ========== Sheet 3: 全部文章 ==========
ws3 = wb.create_sheet('全部文章')
ws3.sheet_properties.tabColor = 'F59E0B'

ws3.merge_cells('A1:G1')
ws3['A1'] = '所有检索到的文章完整列表'
ws3['A1'].font = Font(name='Arial', bold=True, size=16, color='92400E')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 40

row = 2
for i, h in enumerate(['序号', '发布日期', '文章标题', '内容摘要', '主题分类', '发布时段', '星期'], 1):
    cell = ws3.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = PatternFill('solid', fgColor='F59E0B')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for idx, a in enumerate(all_axing, 1):
    row += 1
    dt = a['datetime']
    h = dt.hour
    if 6 <= h < 10: slot = '早间'
    elif 10 <= h < 14: slot = '午间'
    elif 14 <= h < 18: slot = '下午'
    elif 18 <= h < 22: slot = '晚间'
    else: slot = '深夜'
    topic = classify_topic(a['title'], a.get('digest', ''))
    data = [idx, dt.strftime('%Y-%m-%d %H:%M'), a['title'], a.get('digest', '')[:100], topic, slot, weekday_map[dt.weekday()]]
    fill = light_gray_fill if idx % 2 == 0 else white_fill
    for i, v in enumerate(data, 1):
        cell = ws3.cell(row=row, column=i, value=v)
        cell.font = normal_font
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.fill = fill
        cell.border = thin_border

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 18
ws3.column_dimensions['C'].width = 50
ws3.column_dimensions['D'].width = 55
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 10
ws3.column_dimensions['G'].width = 8

# ========== Sheet 4: 分析总结 ==========
ws4 = wb.create_sheet('分析总结')
ws4.sheet_properties.tabColor = '8B5CF6'

ws4.merge_cells('A1:D1')
ws4['A1'] = '数据分析总结与洞察'
ws4['A1'].font = Font(name='Arial', bold=True, size=16, color='5B21B6')
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 40

insights = [
    ('发文频率',
     f'本周（3/12-3/19）共发布 {len(recent_week)} 篇文章，平均每 {7/max(len(recent_week),1):.1f} 天一篇。近30天共发布 {len(recent_month)} 篇，保持较高活跃度。'),
    ('内容定位',
     '账号定位清晰——"AI 学习 & 编程"，作者自称腾讯程序员，内容面向技术爱好者和开发者。从教程到行业观察，涵盖面广。'),
    ('热门话题',
     'OpenClaw（AI Agent 平台）和 MCP 协议是本周核心话题。OpenClaw 相关文章占比最高，已形成系列化运营（"跟阿星一起玩AI"系列）。'),
    ('发布规律',
     '发文时间集中在早间(8-9点)和晚间(20点)。早间适合通勤阅读，晚间适合深度技术文。整体时间策略较成熟。'),
    ('内容趋势',
     '从近期文章看，内容正从"AI 工具介绍"向"AI Agent 实操教程"转型。安全话题（22万实例暴露）表现突出，可作为差异化方向。'),
    ('标题策略',
     '善用疑问句（"为什么你的下一个AI助手应该是OpenClaw?"）、数字（"10个Skill"、"22万实例"）和热词（"裸奔"）制造吸引力。'),
    ('优化建议',
     '1) 可增加周末发文覆盖\n2) 可尝试下午时段发文分流\n3) 加强系列化运营编号\n4) 安全类深度文潜力大，可增加选题\n5) 可增加互动型内容（投票、问答）'),
]

row = 3
for title, content in insights:
    ws4.merge_cells(f'A{row}:D{row}')
    cell = ws4.cell(row=row, column=1, value=title)
    cell.font = Font(name='Arial', bold=True, size=13, color='5B21B6')
    cell.fill = PatternFill('solid', fgColor='EDE9FE')
    cell.border = thin_border
    ws4.row_dimensions[row].height = 30
    row += 1
    ws4.merge_cells(f'A{row}:D{row}')
    cell = ws4.cell(row=row, column=1, value=content)
    cell.font = Font(name='Arial', size=11)
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border
    ws4.row_dimensions[row].height = max(50, content.count('\n') * 20 + 40)
    row += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 25

output_path = '/data/github/openclaw101/跟阿星一起学AI_公众号数据分析.xlsx'
wb.save(output_path)
print(f'报告已生成: {output_path}')
print(f'本周文章数: {len(recent_week)}')
print(f'近30天文章数: {len(recent_month)}')
print(f'总检索文章数: {len(all_axing)}')
